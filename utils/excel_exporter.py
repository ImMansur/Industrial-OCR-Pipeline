import pandas as pd
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    @staticmethod
    def export(data, output_path):
        """Exports extracted data to an Excel file with multiple sheets."""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Certificate Details (Header Information)
            if "structured_info" in data:
                df_info = pd.DataFrame([data["structured_info"]])
                df_info.to_excel(writer, sheet_name="Certificate Details", index=False)
                ExcelExporter._format_sheet(writer.sheets["Certificate Details"])

            # Sheet 2: Equipment Line Items (from tables)
            if "tables" in data and data["tables"]:
                all_line_items = []
                for i, table in enumerate(data["tables"]):
                    df_table = pd.DataFrame(table)
                    if not df_table.empty:
                        # Add certificate reference if available
                        cert_ref = data.get("structured_info", {}).get("Certification Details", f"Table_{i+1}")
                        df_table.insert(0, "Source", cert_ref)
                        all_line_items.append(df_table)
                
                if all_line_items:
                    df_equipment = pd.concat(all_line_items, ignore_index=True)
                    df_equipment.to_excel(writer, sheet_name="Equipment Line Items", index=False)
                    ExcelExporter._format_sheet(writer.sheets["Equipment Line Items"])

            # Sheet 3: Consolidated Register (All parts across certificates)
            if "tables" in data and data["tables"]:
                consolidated_items = []
                for table in data["tables"]:
                    df_table = pd.DataFrame(table)
                    if not df_table.empty:
                        consolidated_items.append(df_table)
                
                if consolidated_items:
                    df_consolidated = pd.concat(consolidated_items, ignore_index=True)
                    # Remove duplicates based on part number if applicable
                    df_consolidated.to_excel(writer, sheet_name="Consolidated Register", index=False)
                    ExcelExporter._format_sheet(writer.sheets["Consolidated Register"])

        print(f"Excel file saved at: {output_path}")

    @staticmethod
    def _format_sheet(worksheet):
        """Apply professional formatting to worksheet"""
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Format data cells and auto-adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top row
        worksheet.freeze_panes = "A2"
