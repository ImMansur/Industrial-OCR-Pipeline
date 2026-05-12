import os
import re
import sys
from utils.azure_ocr import AzureOCR
from utils.azure_openai import OpenAIStructurer
from utils.excel_exporter import ExcelExporter
from dotenv import load_dotenv

load_dotenv()

HEADER_KEYWORDS = {
    "item", "item no", "item no.", "qty", "quantity", "part", "part no", "part number",
    "description", "desc", "serial", "serial no", "serial number", "serialization",
    "pr", "pr level", "notes", "note", "lot", "batch", "expiration", "exp"
}


def merge_header_rows(header_rows):
    """Combine multiple header rows into a single header row."""
    max_cols = max(len(row) for row in header_rows)
    merged = []
    for col_index in range(max_cols):
        parts = [str(row[col_index]).strip() for row in header_rows if col_index < len(row) and str(row[col_index]).strip()]
        merged.append(" ".join(parts).strip())
    return merged


def is_header_row(row):
    """Return True if a row looks like a table header row."""
    text_cells = [str(cell).strip().lower() for cell in row if str(cell).strip()]
    if not text_cells:
        return True
    keyword_count = sum(1 for cell in text_cells if any(keyword in cell for keyword in HEADER_KEYWORDS))
    numeric_count = sum(bool(re.search(r"\d", cell)) for cell in text_cells)
    if keyword_count >= max(1, len(text_cells) // 2):
        return True
    if len(text_cells) == 1 and re.search(r"\b(applicable|expiration|lot|batch|no|n/a|serialization)\b", text_cells[0]):
        return True
    if numeric_count == 0 and len(text_cells) > 1:
        return True
    return False


def clean_table_rows(table):
    """Clean table data by removing header rows and empty rows."""
    if not table or len(table) < 2:
        return [], []

    header_rows = [table[0]]
    row_index = 1
    while row_index < len(table) and is_header_row(table[row_index]):
        header_rows.append(table[row_index])
        row_index += 1

    headers = merge_header_rows(header_rows)
    data_rows = table[row_index:]

    cleaned = []
    for row in data_rows:
        normalized_row = [str(cell).strip() for cell in row]
        if not any(normalized_row):
            continue
        # Skip continuation/comment rows that have no item or quantity
        if len(normalized_row) >= 2 and not normalized_row[0] and not normalized_row[1]:
            continue
        if len(normalized_row) >= 3 and not normalized_row[0] and not normalized_row[2] and not normalized_row[1].isdigit():
            continue
        if len(normalized_row) >= 3 and not normalized_row[1] and not normalized_row[2] and not normalized_row[0].isdigit():
            continue
        if len(normalized_row) >= 3 and normalized_row[0] and not normalized_row[0].isdigit() and not normalized_row[1] and not normalized_row[2]:
            continue
        if len(normalized_row) >= 2 and normalized_row[0] and not normalized_row[0].isdigit() and all(not cell for cell in normalized_row[1:]):
            continue
        # Skip single-cell comment rows
        if len(normalized_row) == 1 and any(keyword in normalized_row[0].lower() for keyword in ['concentration', 'fraction', 'sulphur', 'ph', 'mass', 'elemental', 'chloride']):
            continue
        if all(any(keyword in cell.lower() for keyword in HEADER_KEYWORDS) for cell in normalized_row if cell):
            continue
        cleaned.append(normalized_row[:len(headers)] + [""] * max(0, len(headers) - len(normalized_row)))

    return headers, cleaned

def normalize_table_columns(df_table):
    """Normalize table columns to standard format."""
    # Create case-insensitive mapping
    column_mapping = {}
    standard_cols = {
        "item no": "Item No",
        "quantity": "Quantity", 
        "part number": "Part Number",
        "description": "Description",
        "serial number": "Serial Number",
        "pr level": "PR Level",
        "notes": "Notes",
        # Variations
        "part no": "Part Number",
        "serial no": "Serial Number", 
        "item": "Item No",
        "qty": "Quantity",
        "desc": "Description",
        "part": "Part Number",
        "serial": "Serial Number",
        "note": "Notes",
        "pr": "PR Level",
        "serialization": "Serial Number",
        "lot / batch no. (if applicable)": "Notes",
        "expiration date (if applicable)": "Notes",
        "s/o / lot & batch / exp.": "Notes"
    }
    
    for col in df_table.columns:
        col_lower = re.sub(r'[^a-z0-9 ]+', ' ', str(col).lower().strip())
        for keyword, standard_name in standard_cols.items():
            if keyword in col_lower:
                column_mapping[col] = standard_name
                break
    
    df_table = df_table.rename(columns=column_mapping)
    
    # Standard columns for equipment line items
    standard_columns = ["Item No", "Quantity", "Part Number", "Description", "Serial Number", "PR Level", "Notes"]
    
    # Add missing columns with "N/A"
    for col in standard_columns:
        if col not in df_table.columns:
            df_table[col] = "N/A"
    
    # Reorder to standard columns
    df_table = df_table[standard_columns]
    
    # Fill any NaN with "N/A"
    df_table = df_table.fillna("N/A")
    
    return df_table


def extract_equipment_summary(tables):
    """Aggregate equipment descriptions, part numbers, and serial numbers from extracted tables."""
    import pandas as pd

    all_rows = []
    for table in tables:
        headers, data_rows = clean_table_rows(table)
        if not headers or not data_rows:
            continue
        df_table = pd.DataFrame(data_rows, columns=headers)
        df_table = normalize_table_columns(df_table)
        df_table = df_table[["Description", "Part Number", "Serial Number"]]
        df_table = df_table.drop_duplicates()
        df_table = df_table.loc[df_table.apply(lambda row: any(str(row[col]).strip() and str(row[col]).strip() != "N/A" for col in df_table.columns), axis=1)]
        if not df_table.empty:
            all_rows.append(df_table)

    if not all_rows:
        return None, None, None

    df_all = pd.concat(all_rows, ignore_index=True).drop_duplicates()
    descriptions = [str(v).strip() for v in df_all["Description"].tolist() if str(v).strip() and str(v).strip() != "N/A"]
    parts = [str(v).strip() for v in df_all["Part Number"].tolist() if str(v).strip() and str(v).strip() != "N/A"]
    serials = [str(v).strip() for v in df_all["Serial Number"].tolist() if str(v).strip() and str(v).strip() != "N/A"]

    return (
        ", ".join(dict.fromkeys(descriptions)) or None,
        ", ".join(dict.fromkeys(parts)) or None,
        ", ".join(dict.fromkeys(serials)) or None
    )

def process_document(file_path):
    print(f"--- Processing: {os.path.basename(file_path)} ---")
    
    # 1. OCR Extraction
    ocr = AzureOCR()
    print("Extracting data via Azure Document Intelligence...")
    raw_result = ocr.analyze_document(file_path)
    extracted_data = ocr.extract_data(raw_result)
    
    # 2. LLM Structuring
    structurer = OpenAIStructurer()
    print("Structuring text via Azure OpenAI...")
    structured_info = structurer.structure_text(extracted_data["content"])
    
    # Add metadata
    structured_info["Filename"] = os.path.basename(file_path)
    structured_info["Source Type"] = os.path.splitext(file_path)[1].upper().replace('.', '')
    structured_info["Extraction Status"] = "Success"

    # Extract equipment summary from tables and preserve real values
    equipment_description, part_numbers, serial_numbers = extract_equipment_summary(extracted_data.get("tables", []))
    if equipment_description:
        structured_info["Equipment Or Description"] = equipment_description
    if part_numbers:
        structured_info["Part Numbers"] = part_numbers
    if serial_numbers:
        structured_info["Serial Or Lot Numbers"] = serial_numbers

    extracted_data["structured_info"] = structured_info
    
    # 3. Individual Excel Export (Optional/Backup)
    output_filename = os.path.splitext(os.path.basename(file_path))[0] + "_extracted.xlsx"
    output_path = os.path.join("data", "output", output_filename)
    
    exporter = ExcelExporter()
    exporter.export(extracted_data, output_path)
    print(f"--- Finished: {output_filename} ---\n")
    
    return {
        "structured_info": structured_info,
        "tables": extracted_data.get("tables", []),
        "filename": os.path.basename(file_path)
    }

def main():
    input_dir = os.path.join("data", "input")
    files = [f for f in os.listdir(input_dir) if f.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg'))]
    
    if not files:
        print(f"No documents found in {input_dir}. Please add some files to process.")
        return

    all_results = []
    for file in files:
        file_path = os.path.join(input_dir, file)
        try:
            result = process_document(file_path)
            all_results.append(result)
        except Exception as e:
            print(f"Error processing {file}: {e}")

    # Generate Consolidated Report with 3 Sheets
    if all_results:
        import pandas as pd
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        consolidated_path = os.path.join("data", "output", "consolidated_ocr_report_fixed.xlsx")
        
        try:
            with pd.ExcelWriter(consolidated_path, engine='openpyxl') as writer:
                # Sheet 1: Certificate Details (Header Information from all certificates)
                cert_details = [result["structured_info"] for result in all_results]
                df_certificates = pd.DataFrame(cert_details)
                # Standardize columns to match reviewer sheet
                required_columns = [
                    "Source File",
                    "Source Type", 
                    "Extraction Status",
                    "Customer",
                    "Customer Location",
                    "Sales Order",
                    "Customer Purchase Order",
                    "Job / Project",
                    "Equipment Or Description",
                    "Part Numbers",
                    "Serial Or Lot Numbers",
                    "Certificate Date",
                    "Tested Date",
                    "Lifecycle Date Used",
                    "Recertification Due Date",
                    "Age Months",
                "Months To Recertification",
                "Status",
                "Sales Lead Priority",
                "Invoice Basis",
                "Recommendation",
                "Confidence",
                "Notes",
                "Text Preview"
            ]
            
            # Rename existing columns to match required ones
            column_mapping = {
                "Filename": "Source File",
                "Customer Name": "Customer",
                "Sales Order": "Sales Order",
                "Purchase Order": "Customer Purchase Order",
                "Serialization": "Serial Or Lot Numbers",
                "Certificate Date": "Certificate Date",
                "Document Type": "Equipment Or Description",
                "Applicable Specs": "Job / Project",
                "Address": "Customer Location"
            }
            
            df_certificates = df_certificates.rename(columns=column_mapping)
            
            # Add missing columns with "N/A"
            for col in required_columns:
                if col not in df_certificates.columns:
                    df_certificates[col] = "N/A"
            
            # Reorder columns exactly as specified
            df_certificates = df_certificates[required_columns]
            
            # Replace any NaN values with "N/A"
            df_certificates = df_certificates.fillna("N/A")
            
            df_certificates.to_excel(writer, sheet_name="Certificate Details", index=False)
            
            # Sheet 2: Equipment Line Items (All items from all certificates)
            all_equipment = []
            for result in all_results:
                filename = result["filename"]
                cert_id = result["structured_info"].get("Equipment Serial Number", 
                         result["structured_info"].get("Certification Details", filename))
                
                for table in result.get("tables", []):
                    headers, data_rows = clean_table_rows(table)
                    if not data_rows:
                        continue
                    
                    # Create DataFrame with proper headers
                    df_table = pd.DataFrame(data_rows, columns=headers)
                    
                    # Debug: Log extracted table before normalization
                    print(f"DEBUG - Extracted table for {filename} (first 3 rows):")
                    print(df_table.head(3))
                    
                    # Normalize columns
                    df_table = normalize_table_columns(df_table)
                    
                    # Debug: Log after normalization
                    print(f"DEBUG - Normalized table for {filename} (first 3 rows):")
                    print(df_table.head(3))
                    
                    # Add source references
                    df_table.insert(0, "Certificate", cert_id)
                    df_table.insert(1, "Source File", filename)
                    
                    all_equipment.append(df_table)
            
            if all_equipment:
                df_equipment = pd.concat(all_equipment, ignore_index=True)
                # Remove duplicate rows
                df_equipment = df_equipment.drop_duplicates()
                df_equipment.to_excel(writer, sheet_name="Equipment Line Items", index=False)
            
            # Sheet 3: Consolidated Register (All parts with certificate references)
            if all_equipment:
                df_consolidated = pd.concat([df for df in all_equipment], ignore_index=True)
                # Remove duplicates across all data
                df_consolidated = df_consolidated.drop_duplicates()
                # Sort by Part Number if available
                if "Part Number" in df_consolidated.columns:
                    try:
                        df_consolidated = df_consolidated.sort_values(by="Part Number")
                    except:
                        pass
                df_consolidated.to_excel(writer, sheet_name="Consolidated Register", index=False)
            
            # Format all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                # Apply formatting
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format headers
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border
                
                # Format data cells
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                worksheet.freeze_panes = "A2"
        
            print(f"\nCONSOLIDATED REPORT GENERATED: {consolidated_path}")
            print(f"  ✓ Sheet 1: Certificate Details ({len(cert_details)} certificates)")
            print(f"  ✓ Sheet 2: Equipment Line Items")
            print(f"  ✓ Sheet 3: Consolidated Register")
        except PermissionError:
            print(f"\nWARNING: Could not create consolidated report at {consolidated_path}")
            print("  The file may be open in another application. Individual reports were created successfully.")
        except Exception as e:
            print(f"\nERROR creating consolidated report: {e}")
            print("  Individual reports were created successfully.")

if __name__ == "__main__":
    main()
